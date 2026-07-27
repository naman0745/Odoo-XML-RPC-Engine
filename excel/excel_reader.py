import openpyxl

class ExcelReader:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        self.sheet = self.workbook.active

    def get_headers(self):
        """
        Reads the first row of the active worksheet and returns them as a list of headers.
        """
        headers = []
        for cell in self.sheet[1]:
            headers.append(cell.value)
        return headers

    def get_rows(self):
        """
        Reads all rows after the header and returns them as a list of dictionaries.
        Each dictionary maps the header name to the cell value.
        """
        headers = self.get_headers()
        rows = []

        # iter_rows starting from the second row
        for row_number, row in enumerate(
            self.sheet.iter_rows(min_row=3, values_only=True),
            start=3
        ):
            if all(cell is None for cell in row):
                continue

    # Create a dictionary mapping headers to row values
            row_dict = dict(zip(headers, row))

    # Store the original Excel row number
            row_dict["_row"] = row_number

            rows.append(row_dict)
        return rows

    def close(self):
        """Releases the file handle to prevent OS file locks."""
        if hasattr(self, 'workbook') and self.workbook:
            self.workbook.close()

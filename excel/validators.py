import os
import math
from datetime import date, datetime

import openpyxl

class ExcelValidator:
    def __init__(self, file_path):
        self.file_path = file_path

    def validate_file_exists(self):
        """Checks if the Excel file exists on disk."""
        if not os.path.exists(self.file_path):
            return False, f"File not found: {self.file_path}"
        return True, "File exists"

    def validate_worksheet_exists(self):
        """Ensures the workbook has an active sheet."""
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True)
            try:
                if workbook.active is None:
                    return False, "No active worksheet found in the workbook"
                return True, "Worksheet exists"
            finally:
                workbook.close()
        except Exception as e:
            return False, f"Error opening workbook: {str(e)}"

    def validate_not_empty(self):
        """Ensures the workbook contains at least one row of data (the header)."""
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True)
            try:
                sheet = workbook.active
                if sheet.max_row < 1:
                    return False, "Workbook is empty"
                return True, "Workbook is not empty"
            finally:
                workbook.close()
        except Exception as e:
            return False, f"Error checking workbook: {str(e)}"

    def validate_no_duplicate_headers(self, headers):
        """Ensures there are no overlapping column names in the headers."""
        seen = set()
        duplicates = []
        for h in headers:
            if h is not None:
                if h in seen:
                    duplicates.append(h)
                seen.add(h)

        if duplicates:
            return False, f"Duplicate headers found: {', '.join(map(str, duplicates))}"
        return True, "No duplicate headers"

    def validate_required_columns(self, headers, required_columns):
        """Ensures all required columns are present in the headers."""
        missing = [col for col in required_columns if col not in headers]
        if missing:
            return False, f"Missing required columns: {', '.join(missing)}"
        return True, "All required columns present"

    def validate_no_empty_required_cells(self, row, required_columns):
        """Checks that critical cells in a specific row are not empty."""
        missing_values = [col for col in required_columns if row.get(col) is None or str(row.get(col)).strip() == ""]
        if missing_values:
            return False, f"Empty required cells found: {', '.join(missing_values)}"
        return True, "No empty required cells"

    def validate_positive_number(self, row, field_name, display_name):
        """Checks that a field is numeric, finite, and greater than zero."""
        ok, value_or_msg = self._parse_number(row.get(field_name), display_name)
        if not ok:
            return False, value_or_msg

        if value_or_msg <= 0:
            return False, f"{display_name} must be greater than zero"

        row[field_name] = value_or_msg
        return True, f"{display_name} is valid"

    def validate_non_negative_number(self, row, field_name, display_name):
        """Checks that a field is numeric, finite, and zero or greater."""
        ok, value_or_msg = self._parse_number(row.get(field_name), display_name)
        if not ok:
            return False, value_or_msg

        if value_or_msg < 0:
            return False, f"{display_name} must be zero or greater"

        row[field_name] = value_or_msg
        return True, f"{display_name} is valid"

    def validate_and_normalize_date(self, row, field_name, display_name):
        """
        Checks that a field contains a valid date and normalizes it to YYYY-MM-DD.
        """
        value = row.get(field_name)

        if isinstance(value, datetime):
            row[field_name] = value.date().isoformat()
            return True, f"{display_name} is valid"

        if isinstance(value, date):
            row[field_name] = value.isoformat()
            return True, f"{display_name} is valid"

        if isinstance(value, str):
            value = value.strip()
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
                try:
                    row[field_name] = datetime.strptime(value, fmt).date().isoformat()
                    return True, f"{display_name} is valid"
                except ValueError:
                    continue

        return False, f"{display_name} must be a valid date"

    def _parse_number(self, value, display_name):
        if isinstance(value, bool):
            return False, f"{display_name} must be a number"

        if isinstance(value, (int, float)):
            parsed = float(value)
        elif isinstance(value, str):
            value = value.strip()
            if value == "":
                return False, f"{display_name} must be a number"
            try:
                parsed = float(value)
            except ValueError:
                return False, f"{display_name} must be a number"
        else:
            return False, f"{display_name} must be a number"

        if not math.isfinite(parsed):
            return False, f"{display_name} must be a finite number"

        return True, parsed
